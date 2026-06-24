"""
Generador de Plantillas de Reconsulta - ENIFARM 2026
=====================================================
Llena TODOS los campos disponibles de la plantilla_reconsulta.xlsx
por cada fila del archivo de Repartición (un .xlsx por empresa).

Mapa completo de celdas:
┌─────────────────────────────────────────────────────────────────────────────┐
│ ENCABEZADO                                                                  │
│   M8   → Fecha de generación (hoy)                                          │
├─────────────────────────────────────────────────────────────────────────────┤
│ ANALISTA / SUPERVISOR                                                        │
│   C11  → Analista                     (Repartición: Analista)               │
│   K11  → Supervisor                   (Repartición: Supervisor)             │
├─────────────────────────────────────────────────────────────────────────────┤
│ INFORMACIÓN MUESTRAL (datos del marco)                                      │
│   B13  → CVE_UNICA                    (Repartición: CVE_UNICA)              │
│   E13  → CLASE                        (Directorio: CLASE)                   │
│   G13  → ESTRATO                      (Directorio: ESTRATO_DIS)             │
│   K13  → Razón Social                 (Directorio: E09)                     │
│   B14  → Teléfono                     (sin datos disponibles → vacío)       │
│   E14  → Correo Electrónico           (Directorio: E21)                     │
│   G14  → Página de Internet           (Directorio: E22)                     │
│   K14  → Nombre Empresa               (Directorio: E08)                     │
├─────────────────────────────────────────────────────────────────────────────┤
│ INFORMACIÓN DE LA EMPRESA (CAPTURA) — pre-llenado desde directorio          │
│   D16  → Nombre Empresa captura       (Directorio: E08)                     │
│   K16  → Razón Social captura         (Directorio: E09)                     │
│   B17  → Teléfono 1                   (sin datos disponibles → vacío)       │
│   D17  → Ext. 1                       (vacío)                               │
│   K17  → Correo Electrónico captura   (Directorio: E21)                     │
│   B18  → Teléfono 2                   (vacío)                               │
│   D18  → Ext. 2                       (vacío)                               │
│   K18  → Página de Internet captura   (Directorio: E22)                     │
├─────────────────────────────────────────────────────────────────────────────┤
│ INFORMACIÓN CAPTURA DEL INFORMANTE — el analista llena durante la llamada   │
│   D20  → Nombre del Informante        (vacío)                               │
│   K20  → Puesto                       (vacío)                               │
│   K21  → Correo del Informante        (vacío)                               │
│   B22  → Teléfono del Informante      (vacío)                               │
│   F22  → Ext. Informante              (vacío)                               │
├─────────────────────────────────────────────────────────────────────────────┤
│ INFORMACIÓN DEL ANALISTA — el analista llena al cerrar la reconsulta        │
│   C24  → Fecha de llamadas            (vacío)                               │
│   K24  → Número de llamadas           (vacío)                               │
│   D25  → Fecha de envío de correos    (vacío)                               │
│   K25  → Número de oficio             (vacío)                               │
│   B26  → Validación                   (vacío)                               │
├─────────────────────────────────────────────────────────────────────────────┤
│ OBSERVACIONES                                                                │
│   A28  → Observaciones de Reconsulta  (vacío — analista)                    │
│   A37  → Observaciones de Retorno     (vacío — analista)                    │
└─────────────────────────────────────────────────────────────────────────────┘

Uso:
    python generar_reconsultas.py

Requisitos:
    pip install pandas openpyxl
"""

import pandas as pd
import openpyxl
import os
from datetime import datetime

# ─── RUTAS DE ARCHIVOS ────────────────────────────────────────────────────────
RUTA_DIRECTORIO  = "Directorio_Enifarm2026_120526_SSCEE.xlsm"
RUTA_REPARTICION = ("Repartición_de_cargas_de_trabajo__primera_parte_"
                    "ENIIFARM_VACIADO__PROD_10062026_0329_15062026.xlsx")
RUTA_PLANTILLA   = "plantilla_reconsulta.xlsx"
CARPETA_SALIDA   = "reconsultas"   # Se crea automáticamente
# ─────────────────────────────────────────────────────────────────────────────


def cargar_datos():
    print("Cargando Directorio...")
    df_dir = pd.read_excel(RUTA_DIRECTORIO, engine="openpyxl")

    print("Cargando Repartición...")
    df_rep = pd.read_excel(RUTA_REPARTICION)

    # Clave de enlace:
    #   Repartición.CLEE (entero 17 dígitos, con ceros a la izquierda)
    #   == Directorio.CLEE[:17]  (primeros 17 chars del CLEE largo)
    df_rep["CLEE_KEY"] = df_rep["CLEE"].astype(str).str.zfill(17)
    df_dir["CLEE_KEY"] = df_dir["CLEE"].astype(str).str[:17]

    dir_lookup = df_dir.set_index("CLEE_KEY").to_dict("index")
    return df_rep, dir_lookup


def v(val):
    """Devuelve cadena vacía si el valor es NaN/None/float-nan."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return val


def llenar_plantilla(row, dir_data):
    wb = openpyxl.load_workbook(RUTA_PLANTILLA)
    ws = wb.active

    # ── ENCABEZADO ───────────────────────────────────────────────────────────
    ws["M8"] = datetime.now().date()        # Fecha de generación

    # ── ANALISTA / SUPERVISOR ─────────────────────────────────────────────────
    ws["C11"] = str(row["Analista"])
    ws["K11"] = str(row["Supervisor"])

    # ── INFORMACIÓN MUESTRAL ──────────────────────────────────────────────────
    ws["B13"] = int(row["CVE_UNICA"])           # CVE_UNICA
    ws["E13"] = v(dir_data.get("CLASE"))         # CLASE
    ws["G13"] = str(v(dir_data.get("ESTRATO_DIS"))).strip()  # ESTRATO
    ws["K13"] = v(dir_data.get("E09"))           # Razón Social
    # B14 → Teléfono muestral: no hay teléfono en el directorio
    ws["E14"] = v(dir_data.get("E21"))           # Correo Electrónico
    ws["G14"] = v(dir_data.get("E22"))           # Página de Internet
    ws["K14"] = v(dir_data.get("E08"))           # Nombre Empresa

    # ── INFORMACIÓN DE LA EMPRESA (CAPTURA) ───────────────────────────────────
    ws["D16"] = v(dir_data.get("E08"))           # Nombre Empresa (captura)
    ws["K16"] = v(dir_data.get("E09"))           # Razón Social   (captura)
    # B17 → Teléfono 1: sin datos
    # D17 → Ext. 1:     sin datos
    ws["K17"] = v(dir_data.get("E21"))           # Correo Electrónico (captura)
    # B18 → Teléfono 2: sin datos
    # D18 → Ext. 2:     sin datos
    ws["K18"] = v(dir_data.get("E22"))           # Página de Internet (captura)

    # ── INFORMACIÓN CAPTURA DEL INFORMANTE ───────────────────────────────────
    # D20 → Nombre Informante  )
    # K20 → Puesto             ) El analista llena durante la llamada
    # K21 → Correo Informante  )
    # B22 → Teléfono           )
    # F22 → Ext.               )

    # ── INFORMACIÓN DEL ANALISTA ─────────────────────────────────────────────
    # C24 → Fecha de llamadas          )
    # K24 → Número de llamadas         )
    # D25 → Fecha envío correos        ) El analista llena al cerrar
    # K25 → Número de oficio           )
    # B26 → Validación                 )

    # ── OBSERVACIONES ────────────────────────────────────────────────────────
    # A28 → Observaciones de Reconsulta  → analista
    # A37 → Observaciones de Retorno     → analista

    return wb


def main():
    os.makedirs(CARPETA_SALIDA, exist_ok=True)

    df_rep, dir_lookup = cargar_datos()
    total     = len(df_rep)
    generados = 0
    errores   = []

    print(f"\nGenerando {total} plantillas de reconsulta...\n")

    for _, row in df_rep.iterrows():
        clee_key = row["CLEE_KEY"]
        dir_data = dir_lookup.get(clee_key, {})

        if not dir_data:
            errores.append(f"ADVERTENCIA: CLEE {clee_key} no encontrado en Directorio (CVE {row['CVE_UNICA']})")

        try:
            wb = llenar_plantilla(row, dir_data)
            nombre = f"{CARPETA_SALIDA}/{int(row['ID_CAT_ENCUESTAS_INFO'])}.xlsx"
            wb.save(nombre)
            generados += 1
        except Exception as e:
            errores.append(f"ERROR en CLEE {clee_key} (CVE {row['CVE_UNICA']}): {e}")

    print(f"✓ Generados: {generados} / {total}")
    if errores:
        print(f"\nAvisos / Errores ({len(errores)}):")
        for msg in errores:
            print(f"  {msg}")
    else:
        print("  Sin errores.")

    print(f"\nArchivos guardados en: ./{CARPETA_SALIDA}/")


if __name__ == "__main__":
    main()
