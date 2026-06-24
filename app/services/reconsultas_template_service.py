from openpyxl import load_workbook
import os
from datetime import datetime
import pandas as pd
from app.services.mapeo_automatico import detectar_campos


def limpiar(val):
    if pd.isna(val):
        return ""
    return str(val)


def generar_reconsulta_template(
    df,
    df_errores,
    ruta_salida,
    plantilla_path
):

    mapa = detectar_campos(df)

    def get(fila, key):
        col = mapa.get(key)
        if col and col in fila:
            return limpiar(fila[col])
        return ""

    # ✅ 🔥 USAR SIEMPRE ID_CAT_ENCUESTAS_INFO
    grupos = df_errores.groupby("ID_CAT_ENCUESTAS_INFO")

    os.makedirs(ruta_salida, exist_ok=True)

    for clave, grupo in grupos:

        # ✅ BUSCAR con ID real
        fila_df = df[df["ID_CAT_ENCUESTAS_INFO"] == clave]

        if fila_df.empty:
            continue

        fila = fila_df.iloc[0]

        wb = load_workbook(plantilla_path)
        ws = wb.active

        hoy = datetime.now()

        # =========================
        # FECHA
        # =========================
        ws.cell(6, 8, hoy.strftime("%d/%m/%Y"))
        ws.cell(6, 9, hoy.strftime("%H:%M"))

        # =========================
        # ANALISTA / SUPERVISOR
        # =========================
        ws.cell(10, 2, get(fila, "ANALISTA"))
        ws.cell(10, 6, get(fila, "SUPERVISOR"))

        # =========================
        # INFORMACIÓN MUESTRAL
        # =========================
        ws.cell(13, 2, get(fila, "CLEE"))

        # ✅ puede venir del directorio
        ws.cell(13, 8, get(fila, "RAZON_SOCIAL"))
        ws.cell(14, 8, get(fila, "NOMBRE_EMPRESA"))

        # ✅ puede venir de cualquiera de los 3 archivos
        ws.cell(14, 2, get(fila, "TELEFONO"))
        ws.cell(14, 4, get(fila, "CORREO"))

        # =========================
        # INFORMANTE
        # =========================
        ws.cell(18, 2, get(fila, "INFORMANTE"))

        # =========================
        # TABLA DE ERRORES
        # =========================
        fila_inicio = 32

        for _, r in grupo.iterrows():

            ws.cell(row=fila_inicio, column=2,
                    value=limpiar(r.get("Variables Involucradas")))

            ws.cell(row=fila_inicio, column=3,
                    value=limpiar(r.get("Nombre Vector")))

            ws.cell(row=fila_inicio, column=4,
                    value=limpiar(r.get("Procedimiento")))

            fila_inicio += 1

        archivo = os.path.join(
            ruta_salida,
            f"Reconsulta_{clave}.xlsx"
        )

        wb.save(archivo)